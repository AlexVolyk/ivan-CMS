from os import path
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from .variables import titles, characters, ext

def create_file(path_name, data):
    wb = Workbook()
    ws = wb.active
    character_index = 0
    for i in titles:
        ws[characters[character_index]+'1'] = titles[i]
        ws[characters[character_index]+'1'].fill = PatternFill(start_color='FFC000', patternType='solid')
        ws[characters[character_index]+'2'] = data[i]
        character_index += 1

    wb.save(path_name)



def create_total_file(directory, total_statistic_obj):
    total_name = 'total_statistic'
    total_path_name = path.join(directory, total_name)
    total_path_name = total_path_name + ext
    create_file(path_name=total_path_name, data=total_statistic_obj)


def create_year_file(year_path_name_directory, i, year_statistic_obj):
    year_path_name = path.join(year_path_name_directory,i)
    year_path_name = '%s%s' % (year_path_name, ext)
    create_file(path_name=year_path_name, data=year_statistic_obj)



def create_month_file(year_path_name_directory, j, month_statistic_obj):
    month_path_name = path.join(year_path_name_directory, j, j)
    month_path_name = "%s%s" % (month_path_name, ext)
    create_file(path_name=month_path_name, data=month_statistic_obj)