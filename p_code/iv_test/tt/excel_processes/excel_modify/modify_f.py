from os import listdir
from openpyxl import load_workbook

from .help_f import get_index_and_is_found, get_letter_with_index, is_row_empty, is_row_not_empty_before_12
from .change_status_state_styling import change_status_state_styling
from .update_row import update_row




def get_create_dict(s_v, data=False):

    # ws[get_letter_with_index('A', what_index)] = what_looking_for
    # ws[get_letter_with_index('B', what_index)] = d['type']
    # ws[get_letter_with_index('C', what_index)] = d['brand']
    # # ws[get_letter_with_index('D', what_index)] = d['marked']
    # ws[get_letter_with_index('E', what_index)] = d['color']
    # ws[get_letter_with_index('F', what_index)] = d['notice']
    # ws[get_letter_with_index('G', what_index)] = d['country']
    # ws[get_letter_with_index('H', what_index)] = d['size']
    # ws[get_letter_with_index('I', what_index)] = d['sm']
    # ws[get_letter_with_index('J', what_index)] = d['price']
    # ws[get_letter_with_index('K', what_index)] = d['old_price']
    # # ws[get_letter_with_index('L', what_index)] = d['status']
    # # ws[get_letter_with_index('M', what_index)] = d['state']
    # # ws[get_letter_with_index('N', what_index)] = d['extra_notice']
    if data:
        obj = {
        'A': data['number'],
        'B': data['type'],
        'C': data['brand'],
        'D': data['marked'],
        'E': data['color'],
        'F': data['notice'],
        'G': data['country'],
        'H': data['size'],
        'I': data['sm'],
        'J': data['price'],
        'K': data['old_price'],
        'L': data['status'],
        'M': data['state'],
        'N': data['extra_notice'],
        }
        # print(obj,'objobjobjobjobjobjobjobjobjobjobj')
        return obj

    obj = {
        'A': s_v,
        'B': 'dsad',
        'C': '31',
        'D': '43',
        'E': 'dsa',
        'F': 'gf',
        'G': 'nb',
        'H': '875',
        'I': '065',
        'J': '.nb',
        'K': '.fh2',
        'L': 'Продано',
        'M': 'Пошкоджений',
        'N': '----'
    }
    return obj


def remove_f(name, what_looking_for=False):
    # print(name, what_looking_for)
    if what_looking_for:
        wb = load_workbook(name)

        ws = wb.active
        A = ws['A']


        what_index, is_found = get_index_and_is_found(what_looking_for, A)

        if is_found:
            what_index_2, is_not_empty = is_row_not_empty_before_12(what_looking_for, A)
            if is_not_empty:
                data = {
                'number': '',
                'type': '',
                'brand': '',
                'marked': '',
                'color': '',
                'notice': '',
                'country': '',
                'size': '',
                'sm': '',
                'price': '',
                'old_price': '',
                'status': '',
                'state': '',
                'extra_notice': '',
                'date': ''
                }
                update_row(ws, data, what_index=what_index_2, what_looking_for='')
                change_status_state_styling(ws, what_index_2)
            
            else:
                ws.delete_rows(what_index)
                # val = ws.delete_rows(what_index)
                # get_letter_with_index('A', what_index)
                # print(val, 'A'+ str(what_index))

    wb.save(name)
    # raise('errr')



def add_f(name, what_looking_for=False, data=False):
    # print(name,what_looking_for,data)
    # print(listdir('../excel'))
    

    if what_looking_for and data:
        wb = load_workbook(name)

        ws = wb.active
        A = ws['A']

        # is_found = False

        what_index, is_found = get_index_and_is_found(what_looking_for, A)

        what_index_empty, is_empty = is_row_empty(A)


        if not is_found:
            # print('not found')
            if is_empty:
                update_row(ws, data, what_index_empty, what_looking_for)

                what_looking_for_2 = what_looking_for
                A_2 = ws['A']
                what_index_2, is_found_2 = get_index_and_is_found(what_looking_for_2, A_2)
                if is_found_2:
                    change_status_state_styling(ws, what_index_2)

            else:

                ws.append(get_create_dict(s_v=None, data=data))

                what_looking_for_2 = what_looking_for
                A_2 = ws['A']
                what_index_2, is_found_2 = get_index_and_is_found(what_looking_for_2, A_2)
                if is_found_2:
                    change_status_state_styling(ws, what_index_2)

        elif is_found:
            # print('found_and_update')
            update_row(ws, data, what_index, what_looking_for)

            what_looking_for_2 = what_looking_for
            A_2 = ws['A']
            what_index_2, is_found_2 = get_index_and_is_found(what_looking_for_2, A_2)
            if is_found_2:
                change_status_state_styling(ws, what_index_2)
        
        wb.save(name)
        # raise Exception("Sorry, bruh. you are excuse")

