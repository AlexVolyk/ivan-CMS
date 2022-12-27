from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font

from .style_constants import product_evaluation_styles, sold_styles, unsold_styles

def get_product_evaluation_styles(value):
    value = str(value).lower()
    bg_style = PatternFill('solid')
    color_style = Font()
    is_found = False

    if value in product_evaluation_styles:
        bg_style.start_color = product_evaluation_styles[value]['bg']
        color_style.color = product_evaluation_styles[value]['color']
        is_found = True
    else:
        bg_style.fill_type = None
        color_style.color = '000000'
        is_found = True



    return bg_style, color_style, is_found




def get_sold_or_not_styles(value):

    bg_style = PatternFill('solid')
    color_style = Font()

    is_found = False
    if value == 'Продано':
        bg_style.start_color = sold_styles['bg']
        color_style.color = sold_styles['color']
        is_found = True


    elif value == 'Не продано':
        bg_style.start_color = unsold_styles['bg']
        color_style.color = unsold_styles['color']
        is_found = True

    else:
        bg_style.fill_type = None
        color_style.color = '000000'
        is_found = True



    return bg_style, color_style, is_found